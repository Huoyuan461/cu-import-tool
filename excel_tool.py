from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
import json
import math
from pathlib import Path
import re
import shutil
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet


CONFIG_PATH = Path("rules_config.json")
ALLOWED_OPERATIONS = {"copy_value", "sum_column"}


@dataclass
class ProductRule:
    product_name: str
    customer_value: str = ""
    part_no_value: str = ""
    project_no_value: str = ""


@dataclass
class MappingRule:
    rule_name: str
    product_name: str
    source_sheet: str
    customer_column: str
    part_no_column: str
    project_no_column: str
    operation: str
    source_value_column: str
    target_sheet: str
    target_cell: str
    empty_policy: str = "skip"
    header_row: int = 1


@dataclass
class SourceWorkbook:
    name: str
    data: bytes


@dataclass
class ProcessingLog:
    level: str
    rule_name: str
    product_name: str
    message: str
    source_file: str = ""
    source_sheet: str = ""
    target_sheet: str = ""
    target_cell: str = ""
    value: Any = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "level": self.level,
            "rule_name": self.rule_name,
            "product_name": self.product_name,
            "source_file": self.source_file,
            "source_sheet": self.source_sheet,
            "target_sheet": self.target_sheet,
            "target_cell": self.target_cell,
            "value": self.value,
            "message": self.message,
        }


@dataclass
class RuleResult:
    mapping: MappingRule
    value: Any = None
    matched_count: int = 0
    logs: list[ProcessingLog] = field(default_factory=list)
    should_write: bool = False


def normalize_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def split_keys(value: Any) -> set[str]:
    text = str(value or "").strip()
    if not text:
        return set()
    parts = re.split(r"[,，;；\n]+", text)
    return {normalize_key(part) for part in parts if normalize_key(part)}


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_product_rules(rows: Iterable[dict[str, Any]]) -> list[ProductRule]:
    rules: list[ProductRule] = []
    for row in rows:
        product_name = str(row.get("product_name") or "").strip()
        if not product_name:
            continue
        rules.append(
            ProductRule(
                product_name=product_name,
                customer_value=str(row.get("customer_value") or "").strip(),
                part_no_value=str(row.get("part_no_value") or "").strip(),
                project_no_value=str(row.get("project_no_value") or "").strip(),
            )
        )
    return rules


def parse_mapping_rules(rows: Iterable[dict[str, Any]]) -> list[MappingRule]:
    mappings: list[MappingRule] = []
    for row in rows:
        rule_name = str(row.get("rule_name") or "").strip()
        if not rule_name:
            continue
        header_row = row.get("header_row", 1)
        try:
            header_row_int = int(header_row or 1)
        except ValueError:
            header_row_int = 1
        mappings.append(
            MappingRule(
                rule_name=rule_name,
                product_name=str(row.get("product_name") or "").strip(),
                source_sheet=str(row.get("source_sheet") or "").strip(),
                customer_column=str(row.get("customer_column") or "").strip(),
                part_no_column=str(row.get("part_no_column") or "").strip(),
                project_no_column=str(row.get("project_no_column") or "").strip(),
                operation=str(row.get("operation") or "copy_value").strip(),
                source_value_column=str(row.get("source_value_column") or "").strip(),
                target_sheet=str(row.get("target_sheet") or "").strip(),
                target_cell=str(row.get("target_cell") or "").strip(),
                empty_policy=str(row.get("empty_policy") or "skip").strip(),
                header_row=max(1, header_row_int),
            )
        )
    return mappings


def default_product_rows() -> list[dict[str, str]]:
    return [
        {
            "product_name": "产品A",
            "customer_value": "客户名称",
            "part_no_value": "零件号",
            "project_no_value": "项目号",
        }
    ]


def default_mapping_rows() -> list[dict[str, Any]]:
    return [
        {
            "rule_name": "示例_取数",
            "product_name": "产品A",
            "source_sheet": "Sheet1",
            "customer_column": "客户",
            "part_no_column": "零件号",
            "project_no_column": "项目号",
            "operation": "copy_value",
            "source_value_column": "金额",
            "target_sheet": "Sheet1",
            "target_cell": "B2",
            "empty_policy": "skip",
            "header_row": 1,
        }
    ]


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    if not path.exists():
        return {"products": default_product_rows(), "mappings": default_mapping_rows()}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    data.setdefault("products", default_product_rows())
    data.setdefault("mappings", default_mapping_rows())
    return data


def save_config(products: list[dict[str, Any]], mappings: list[dict[str, Any]], path: Path = CONFIG_PATH) -> None:
    data = {"products": products, "mappings": mappings}
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def resolve_column(ws: Worksheet, column_ref: str, header_row: int = 1) -> int:
    ref = str(column_ref or "").strip()
    if not ref:
        raise ValueError("列配置为空")
    if ref.isdigit():
        index = int(ref)
        if index <= 0:
            raise ValueError(f"列号必须大于0: {ref}")
        return index
    if re.fullmatch(r"[A-Za-z]{1,3}", ref):
        return column_index_from_string(ref.upper())
    wanted = normalize_key(ref)
    for col in range(1, ws.max_column + 1):
        if normalize_key(ws.cell(header_row, col).value) == wanted:
            return col
    raise ValueError(f"找不到表头列: {ref}")


def validate_target_cell(cell_ref: str) -> tuple[int, int]:
    try:
        return coordinate_to_tuple(str(cell_ref or "").strip().upper())
    except Exception as exc:
        raise ValueError(f"目标单元格无效: {cell_ref}") from exc


def product_matches(ws: Worksheet, row: int, product: ProductRule, mapping: MappingRule) -> bool:
    checks = [
        (mapping.customer_column, split_keys(product.customer_value)),
        (mapping.part_no_column, split_keys(product.part_no_value)),
        (mapping.project_no_column, split_keys(product.project_no_value)),
    ]
    for column_ref, accepted_values in checks:
        if not column_ref or not accepted_values:
            continue
        try:
            column_index = resolve_column(ws, column_ref, mapping.header_row)
        except ValueError:
            continue
        if normalize_key(ws.cell(row, column_index).value) in accepted_values:
            return True
    return False


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def prepare_target(target_path: Path) -> Path:
    if not target_path.exists():
        raise FileNotFoundError(f"目标文件不存在: {target_path}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = target_path.with_name(f"{target_path.stem}.backup_{timestamp}{target_path.suffix}")
    shutil.copy2(target_path, backup)
    return backup


def evaluate_rule(
    mapping: MappingRule,
    product: ProductRule | None,
    sources: list[SourceWorkbook],
) -> RuleResult:
    result = RuleResult(mapping=mapping)
    if product is None:
        result.logs.append(
            ProcessingLog("ERROR", mapping.rule_name, mapping.product_name, "找不到对应的产品规则")
        )
        return result
    if mapping.operation not in ALLOWED_OPERATIONS:
        result.logs.append(
            ProcessingLog("ERROR", mapping.rule_name, product.product_name, f"不支持的操作: {mapping.operation}")
        )
        return result

    copied_value: Any = None
    sum_value = 0.0
    sum_seen = False

    for source in sources:
        try:
            wb = load_workbook(BytesIO(source.data), read_only=True, data_only=True)
        except Exception as exc:
            result.logs.append(
                ProcessingLog("ERROR", mapping.rule_name, product.product_name, f"源文件打开失败: {exc}", source_file=source.name)
            )
            continue
        if mapping.source_sheet not in wb.sheetnames:
            result.logs.append(
                ProcessingLog("ERROR", mapping.rule_name, product.product_name, "源sheet不存在", source_file=source.name, source_sheet=mapping.source_sheet)
            )
            wb.close()
            continue
        ws = wb[mapping.source_sheet]
        try:
            value_col = resolve_column(ws, mapping.source_value_column, mapping.header_row)
        except ValueError as exc:
            result.logs.append(
                ProcessingLog("ERROR", mapping.rule_name, product.product_name, str(exc), source_file=source.name, source_sheet=mapping.source_sheet)
            )
            wb.close()
            continue

        for row in range(mapping.header_row + 1, ws.max_row + 1):
            if not product_matches(ws, row, product, mapping):
                continue
            result.matched_count += 1
            raw_value = ws.cell(row, value_col).value
            if mapping.operation == "copy_value":
                if copied_value is None and not is_blank(raw_value):
                    copied_value = raw_value
                    result.logs.append(
                        ProcessingLog(
                            "INFO",
                            mapping.rule_name,
                            product.product_name,
                            f"命中并取第一个非空值，源行 {row}",
                            source_file=source.name,
                            source_sheet=mapping.source_sheet,
                            value=raw_value,
                        )
                    )
            else:
                number = as_number(raw_value)
                if number is None:
                    if not is_blank(raw_value):
                        result.logs.append(
                            ProcessingLog(
                                "WARN",
                                mapping.rule_name,
                                product.product_name,
                                f"求和时忽略非数字值，源行 {row}",
                                source_file=source.name,
                                source_sheet=mapping.source_sheet,
                                value=raw_value,
                            )
                        )
                    continue
                sum_value += number
                sum_seen = True
        wb.close()

    if result.matched_count == 0:
        result.logs.append(
            ProcessingLog("WARN", mapping.rule_name, product.product_name, "未匹配到产品数据")
        )
        return result

    if mapping.operation == "copy_value":
        if copied_value is None:
            result.logs.append(
                ProcessingLog("WARN", mapping.rule_name, product.product_name, "匹配到产品，但取数字段为空")
            )
            return result
        result.value = copied_value
    else:
        if not sum_seen:
            result.logs.append(
                ProcessingLog("WARN", mapping.rule_name, product.product_name, "匹配到产品，但没有可求和的数字")
            )
            return result
        result.value = int(sum_value) if float(sum_value).is_integer() else sum_value

    result.should_write = True
    result.logs.append(
        ProcessingLog("INFO", mapping.rule_name, product.product_name, "规则计算完成", value=result.value)
    )
    return result


def evaluate_rules(
    products: list[ProductRule],
    mappings: list[MappingRule],
    sources: list[SourceWorkbook],
) -> list[RuleResult]:
    product_by_name = {p.product_name: p for p in products}
    return [evaluate_rule(mapping, product_by_name.get(mapping.product_name), sources) for mapping in mappings]


def write_results_to_target(
    target_path: Path,
    results: list[RuleResult],
    overwrite: bool = True,
) -> tuple[bytes, list[ProcessingLog], Path | None]:
    logs: list[ProcessingLog] = []
    backup_path: Path | None = None
    if overwrite:
        backup_path = prepare_target(target_path)

    wb = load_workbook(target_path)
    for result in results:
        mapping = result.mapping
        if not result.should_write:
            logs.extend(result.logs)
            continue
        if mapping.target_sheet not in wb.sheetnames:
            logs.extend(result.logs)
            logs.append(
                ProcessingLog("ERROR", mapping.rule_name, mapping.product_name, "目标sheet不存在", target_sheet=mapping.target_sheet, target_cell=mapping.target_cell, value=result.value)
            )
            continue
        try:
            validate_target_cell(mapping.target_cell)
        except ValueError as exc:
            logs.extend(result.logs)
            logs.append(
                ProcessingLog("ERROR", mapping.rule_name, mapping.product_name, str(exc), target_sheet=mapping.target_sheet, target_cell=mapping.target_cell, value=result.value)
            )
            continue
        ws = wb[mapping.target_sheet]
        ws[mapping.target_cell] = result.value
        logs.extend(result.logs)
        logs.append(
            ProcessingLog(
                "INFO",
                mapping.rule_name,
                mapping.product_name,
                "已写入目标单元格",
                target_sheet=mapping.target_sheet,
                target_cell=mapping.target_cell,
                value=result.value,
            )
        )

    output = BytesIO()
    wb.save(output)
    wb.close()
    data = output.getvalue()
    if overwrite:
        target_path.write_bytes(data)
    return data, logs, backup_path


def write_results_to_uploaded_target(
    target_bytes: bytes,
    results: list[RuleResult],
) -> tuple[bytes, list[ProcessingLog]]:
    logs: list[ProcessingLog] = []
    wb = load_workbook(BytesIO(target_bytes))
    for result in results:
        mapping = result.mapping
        if not result.should_write:
            logs.extend(result.logs)
            continue
        if mapping.target_sheet not in wb.sheetnames:
            logs.extend(result.logs)
            logs.append(ProcessingLog("ERROR", mapping.rule_name, mapping.product_name, "目标sheet不存在", target_sheet=mapping.target_sheet, target_cell=mapping.target_cell, value=result.value))
            continue
        try:
            validate_target_cell(mapping.target_cell)
        except ValueError as exc:
            logs.extend(result.logs)
            logs.append(ProcessingLog("ERROR", mapping.rule_name, mapping.product_name, str(exc), target_sheet=mapping.target_sheet, target_cell=mapping.target_cell, value=result.value))
            continue
        wb[mapping.target_sheet][mapping.target_cell] = result.value
        logs.extend(result.logs)
        logs.append(ProcessingLog("INFO", mapping.rule_name, mapping.product_name, "已写入目标单元格", target_sheet=mapping.target_sheet, target_cell=mapping.target_cell, value=result.value))
    output = BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue(), logs
