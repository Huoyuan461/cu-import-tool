from __future__ import annotations

from io import BytesIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from excel_tool import (
    MappingRule,
    ProductRule,
    SourceWorkbook,
    evaluate_rules,
    normalize_key,
    write_results_to_target,
)


def workbook_bytes(rows, sheet_name="Data") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def make_target(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Target"
    ws["A1"] = "Keep"
    ws["B1"] = "=1+1"
    ws["B2"] = "old"
    ws["C3"] = 0
    wb.create_sheet("Other")["A1"] = "untouched"
    wb.save(path)


class ExcelToolTests(unittest.TestCase):
    def test_normalize_key_cleans_spaces_case_and_numbers(self):
        self.assertEqual(normalize_key(" abc "), "ABC")
        self.assertEqual(normalize_key("A   B"), "A B")
        self.assertEqual(normalize_key(123.0), "123")

    def test_copy_value_matches_by_any_product_field(self):
        source = SourceWorkbook(
            "source.xlsx",
            workbook_bytes(
                [
                    ["客户", "零件号", "项目号", "金额"],
                    ["Other", "PN-1", "P-1", 10],
                    ["客户A", "PN-2", "P-2", 88],
                ]
            ),
        )
        product = ProductRule("产品A", customer_value="客户A")
        mapping = MappingRule(
            rule_name="copy",
            product_name="产品A",
            source_sheet="Data",
            customer_column="客户",
            part_no_column="零件号",
            project_no_column="项目号",
            operation="copy_value",
            source_value_column="金额",
            target_sheet="Target",
            target_cell="B2",
        )
        result = evaluate_rules([product], [mapping], [source])[0]
        self.assertTrue(result.should_write)
        self.assertEqual(result.value, 88)
        self.assertEqual(result.matched_count, 1)

    def test_sum_column_sums_all_matches_across_multiple_sources(self):
        sources = [
            SourceWorkbook(
                "one.xlsx",
                workbook_bytes(
                    [
                        ["客户", "零件号", "项目号", "金额"],
                        ["客户A", "PN-1", "P-1", 10],
                        ["客户A", "PN-2", "P-2", "20"],
                        ["客户B", "PN-3", "P-3", 99],
                    ]
                ),
            ),
            SourceWorkbook(
                "two.xlsx",
                workbook_bytes(
                    [
                        ["客户", "零件号", "项目号", "金额"],
                        ["客户A", "PN-4", "P-4", 5.5],
                        ["客户A", "PN-5", "P-5", "not-number"],
                    ]
                ),
            ),
        ]
        product = ProductRule("产品A", customer_value="客户A")
        mapping = MappingRule(
            rule_name="sum",
            product_name="产品A",
            source_sheet="Data",
            customer_column="客户",
            part_no_column="零件号",
            project_no_column="项目号",
            operation="sum_column",
            source_value_column="金额",
            target_sheet="Target",
            target_cell="C3",
        )
        result = evaluate_rules([product], [mapping], sources)[0]
        self.assertTrue(result.should_write)
        self.assertEqual(result.value, 35.5)
        self.assertEqual(result.matched_count, 4)
        self.assertTrue(any(log.level == "WARN" for log in result.logs))

    def test_write_results_overwrites_target_and_creates_backup_preserving_formula(self):
        source = SourceWorkbook(
            "source.xlsx",
            workbook_bytes([["客户", "零件号", "项目号", "金额"], ["客户A", "PN", "P", 123]])
        )
        product = ProductRule("产品A", customer_value="客户A")
        mapping = MappingRule(
            rule_name="copy",
            product_name="产品A",
            source_sheet="Data",
            customer_column="客户",
            part_no_column="零件号",
            project_no_column="项目号",
            operation="copy_value",
            source_value_column="金额",
            target_sheet="Target",
            target_cell="B2",
        )
        result = evaluate_rules([product], [mapping], [source])
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "target.xlsx"
            make_target(target)
            _, logs, backup = write_results_to_target(target, result, overwrite=True)
            self.assertIsNotNone(backup)
            self.assertTrue(backup.exists())
            wb = load_workbook(target, data_only=False)
            self.assertEqual(wb["Target"]["B2"].value, 123)
            self.assertEqual(wb["Target"]["B1"].value, "=1+1")
            self.assertEqual(wb["Other"]["A1"].value, "untouched")
            self.assertTrue(any(log.message == "已写入目标单元格" for log in logs))

    def test_missing_sheet_is_logged_and_not_written(self):
        source = SourceWorkbook("source.xlsx", workbook_bytes([["客户", "金额"], ["客户A", 1]]))
        product = ProductRule("产品A", customer_value="客户A")
        mapping = MappingRule(
            rule_name="bad",
            product_name="产品A",
            source_sheet="Missing",
            customer_column="客户",
            part_no_column="",
            project_no_column="",
            operation="copy_value",
            source_value_column="金额",
            target_sheet="Target",
            target_cell="B2",
        )
        result = evaluate_rules([product], [mapping], [source])[0]
        self.assertFalse(result.should_write)
        self.assertTrue(any("源sheet不存在" in log.message for log in result.logs))


if __name__ == "__main__":
    unittest.main()
